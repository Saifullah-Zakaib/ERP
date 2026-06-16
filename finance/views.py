from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Sum
from datetime import date, datetime
from decimal import Decimal
import json
from .models import Invoice, InvoiceItem, Expense, Payment, FinancialReport, Budget
from core.models import AuditLog
from core.views import role_required, get_client_ip
from hr.models import Payroll
from suppliers.models import PurchaseOrder

def dashboard(request):
    """Finance dashboard"""
    from datetime import timedelta
    from django.utils import timezone
    

    # Get current month data
    current_month = timezone.now().month
    current_year = timezone.now().year
    last_month = current_month - 1 if current_month > 1 else 12
    last_month_year = current_year if current_month > 1 else current_year - 1

    # Calculate total revenue (paid invoices)
    total_revenue = Invoice.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    current_month_revenue = Invoice.objects.filter(
        status='paid',
        invoice_date__month=current_month,
        invoice_date__year=current_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    last_month_revenue = Invoice.objects.filter(
        status='paid',
        invoice_date__month=last_month,
        invoice_date__year=last_month_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # Calculate revenue change percentage
    if last_month_revenue > 0:
        revenue_change = ((current_month_revenue - last_month_revenue) / last_month_revenue) * 100
    else:
        revenue_change = 100 if current_month_revenue > 0 else 0

    # Calculate total expenses from multiple sources
    # 1. Regular expenses from Expense model
    expense_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    expense_current_month = Expense.objects.filter(
        expense_date__month=current_month,
        expense_date__year=current_year
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    expense_last_month = Expense.objects.filter(
        expense_date__month=last_month,
        expense_date__year=last_month_year
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    # 2. Payroll expenses from HR module
    payroll_total = Payroll.objects.aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    payroll_current_month = Payroll.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    payroll_last_month = Payroll.objects.filter(
        payment_date__month=last_month,
        payment_date__year=last_month_year
    ).aggregate(Sum('net_pay'))['net_pay__sum'] or 0

    # 3. Purchase Order expenses from Suppliers module
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    po_current_month = PurchaseOrder.objects.filter(
        status='received',
        order_date__month=current_month,
        order_date__year=current_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    po_last_month = PurchaseOrder.objects.filter(
        status='received',
        order_date__month=last_month,
        order_date__year=last_month_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # Combine all expenses
    total_expenses = expense_total + payroll_total + po_total
    current_month_expenses = expense_current_month + payroll_current_month + po_current_month
    last_month_expenses = expense_last_month + payroll_last_month + po_last_month

    # Calculate expense change percentage
    if last_month_expenses > 0:
        expense_change = ((current_month_expenses - last_month_expenses) / last_month_expenses) * 100
    else:
        expense_change = 100 if current_month_expenses > 0 else 0

    # Calculate net profit
    net_profit = total_revenue - total_expenses
    current_month_profit = current_month_revenue - current_month_expenses
    last_month_profit = last_month_revenue - last_month_expenses

    # Calculate profit change percentage
    if last_month_profit != 0:
        profit_change = ((current_month_profit - last_month_profit) / abs(last_month_profit)) * 100
    else:
        profit_change = 100 if current_month_profit > 0 else 0

    # Pending invoices
    pending_invoices_count = Invoice.objects.filter(status__in=['draft', 'sent']).count()
    pending_invoices_amount = Invoice.objects.filter(status__in=['draft', 'sent']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # Expense breakdown by category - combine all sources
    expense_breakdown = []
    
    # Add regular expenses by category
    regular_expenses = Expense.objects.values('category').annotate(total=Sum('amount'))
    for exp in regular_expenses:
        expense_breakdown.append({
            'category': exp['category'],
            'total': float(exp['total'])
        })
    
    # Add payroll as labor costs
    if payroll_total > 0:
        expense_breakdown.append({
            'category': 'labor_costs',
            'total': float(payroll_total)
        })
    
    # Add purchase orders as raw materials
    if po_total > 0:
        expense_breakdown.append({
            'category': 'raw_materials',
            'total': float(po_total)
        })
    
    # Calculate percentages
    if total_expenses > 0:
        for exp in expense_breakdown:
            exp['percentage'] = (exp['total'] / float(total_expenses)) * 100
    
    # Sort by total descending
    expense_breakdown = sorted(expense_breakdown, key=lambda x: x['total'], reverse=True)
    
    # Recent expenses - combine from all sources
    recent_expenses_list = []
    
    # Add regular expenses
    for exp in Expense.objects.all().order_by('-expense_date')[:5]:
        recent_expenses_list.append({
            'description': exp.description,
            'amount': float(exp.amount),
            'expense_date': exp.expense_date,
            'category': exp.get_category_display()
        })
    
    # Add recent payroll
    for payroll in Payroll.objects.all().order_by('-payment_date')[:3]:
        recent_expenses_list.append({
            'description': f'Payroll - {payroll.employee.first_name} {payroll.employee.last_name}',
            'amount': float(payroll.net_pay),
            'expense_date': payroll.payment_date,
            'category': 'Payroll'
        })
    
    # Add recent purchase orders
    for po in PurchaseOrder.objects.filter(status='received').order_by('-order_date')[:3]:
        recent_expenses_list.append({
            'description': f'Purchase Order - {po.get_supplier_name()}',
            'amount': float(po.total_amount),
            'expense_date': po.order_date,
            'category': 'Purchase Order'
        })
    
    # Sort by date and take top 5
    recent_expenses_list = sorted(recent_expenses_list, key=lambda x: x['expense_date'], reverse=True)[:5]
    
    # Get notifications for current user
    from core.models import SystemNotification
    if request.user.is_authenticated:
        notifications = SystemNotification.objects.filter(
            user=request.user,
            is_read=False
        ).order_by('-created_at')[:10]
        unread_notification_count = SystemNotification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
    else:
        notifications = []
        unread_notification_count = 0
    
    context = {
        'total_revenue': total_revenue / 1000,  # Convert to K
        'revenue_change': round(revenue_change, 1),
        'revenue_change_abs': round(abs(revenue_change), 1),
        'total_expenses': total_expenses / 1000,  # Convert to K
        'expense_change': round(expense_change, 1),
        'expense_change_abs': round(abs(expense_change), 1),
        'net_profit': net_profit / 1000,  # Convert to K
        'profit_change': round(profit_change, 1),
        'profit_change_abs': round(abs(profit_change), 1),
        'pending_invoices_count': pending_invoices_count,
        'pending_invoices_amount': pending_invoices_amount,
        'recent_invoices': Invoice.objects.all().order_by('-invoice_date')[:10],
        'recent_expenses': recent_expenses_list,
        'expense_breakdown': expense_breakdown,
        'notifications': notifications,
        'unread_notification_count': unread_notification_count,
    }

    return render(request, 'finance/dashboard.html', context)



@login_required
def invoice_list(request):
    """List all invoices"""
    invoices = Invoice.objects.all()
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        invoices = invoices.filter(status=status)
    
    context = {
        'invoices': invoices,
        'status_choices': Invoice.STATUS_CHOICES,
    }
    
    return render(request, 'finance/invoice_list.html', context)


@login_required
def create_invoice(request):
    """Create new invoice"""
    if request.method == 'POST':
        invoice = Invoice.objects.create(
            invoice_number=request.POST.get('invoice_number'),
            customer_name=request.POST.get('customer_name'),
            invoice_date=request.POST.get('invoice_date'),
            due_date=request.POST.get('due_date'),
            total_amount=float(request.POST.get('total_amount', 0)),
            notes=request.POST.get('notes', ''),
            created_by=request.user,
        )
        
        # Add invoice items
        item_count = int(request.POST.get('item_count', 0))
        subtotal = 0
        
        for i in range(item_count):
            description = request.POST.get(f'item_description_{i}')
            quantity = int(request.POST.get(f'item_quantity_{i}', 0))
            unit_price = float(request.POST.get(f'item_price_{i}', 0))
            
            if description and quantity and unit_price:
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=description,
                    quantity=quantity,
                    unit_price=unit_price,
                )
                subtotal += quantity * unit_price
        
        invoice.subtotal = subtotal
        invoice.total_amount = subtotal
        invoice.save()
        
        AuditLog.objects.create(
            user=request.user,
            action='Create Invoice',
            module='Finance',
            description=f'Created invoice {invoice.invoice_number}',
        )
        
        messages.success(request, 'Invoice created successfully!')
        return redirect('finance:invoice_detail', invoice_id=invoice.id)
    
    return render(request, 'finance/create_invoice.html')


@login_required
def invoice_detail(request, invoice_id):
    """View invoice details"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = InvoiceItem.objects.filter(invoice=invoice)
    
    context = {
        'invoice': invoice,
        'items': items,
    }
    
    return render(request, 'finance/invoice_detail.html', context)


@login_required
def expense_list(request):
    """List all expenses"""
    expenses = Expense.objects.all()
    
    # Filter by category
    category = request.GET.get('category')
    if category:
        expenses = expenses.filter(category=category)
    
    context = {
        'expenses': expenses,
        'categories': Expense.CATEGORY_CHOICES,
    }
    
    return render(request, 'finance/expense_list.html', context)


@login_required
def add_expense(request):
    """Add new expense"""
    if request.method == 'POST':
        expense = Expense.objects.create(
            expense_number=request.POST.get('expense_number'),
            category=request.POST.get('category'),
            description=request.POST.get('description'),
            amount=float(request.POST.get('amount')),
            expense_date=request.POST.get('expense_date'),
            notes=request.POST.get('notes', ''),
            recorded_by=request.user,
        )
        
        AuditLog.objects.create(
            user=request.user,
            action='Add Expense',
            module='Finance',
            description=f'Added expense {expense.expense_number}',
        )
        
        messages.success(request, 'Expense added successfully!')
        return redirect('finance:expense_list')
    
    return render(request, 'finance/add_expense.html', {
        'categories': Expense.CATEGORY_CHOICES
    })


@login_required
def payment_list(request):
    """List all payments"""
    payments = Payment.objects.all()
    
    context = {
        'payments': payments,
    }
    
    return render(request, 'finance/payment_list.html', context)


@login_required
def record_payment(request):
    """Record new payment"""
    if request.method == 'POST':
        payment = Payment.objects.create(
            payment_number=request.POST.get('payment_number'),
            payment_type=request.POST.get('payment_type'),
            payment_method=request.POST.get('payment_method'),
            amount=float(request.POST.get('amount')),
            payment_date=request.POST.get('payment_date'),
            recipient=request.POST.get('recipient'),
            reference_number=request.POST.get('reference_number', ''),
            notes=request.POST.get('notes', ''),
            processed_by=request.user,
        )
        
        AuditLog.objects.create(
            user=request.user,
            action='Record Payment',
            module='Finance',
            description=f'Recorded payment {payment.payment_number}',
        )
        
        messages.success(request, 'Payment recorded successfully!')
        return redirect('finance:payment_list')
    
    return render(request, 'finance/record_payment.html', {
        'payment_types': Payment.PAYMENT_TYPES,
        'payment_methods': Payment.PAYMENT_METHODS,
    })


@login_required
def reports(request):
    """Generate financial reports"""
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        report_period = request.POST.get('report_period')
        
        report = FinancialReport.objects.create(
            report_type=report_type,
            report_period=report_period,
            generated_by=request.user,
        )
        
        messages.success(request, 'Report generated successfully!')
        return redirect('finance:reports')
    
    reports = FinancialReport.objects.all()[:20]
    return render(request, 'finance/reports.html', {'reports': reports})



# ==================== API ENDPOINTS ====================

@login_required
@role_required('finance', 'admin')
def create_invoice_api(request):
    """Create invoice via API"""
    if request.method == 'POST':
        try:
            # Get form data
            invoice_number = request.POST.get('invoice_number')
            customer_name = request.POST.get('customer_name')
            invoice_date = request.POST.get('invoice_date')
            items_json = request.POST.get('items', '[]')
            
            # Validate required fields
            if not all([invoice_number, customer_name, invoice_date]):
                return JsonResponse({
                    'success': False,
                    'message': 'All fields are required'
                })
            
            # Check if invoice number already exists
            if Invoice.objects.filter(invoice_number=invoice_number).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Invoice number already exists'
                })
            
            # Parse items
            items = json.loads(items_json)
            if not items:
                return JsonResponse({
                    'success': False,
                    'message': 'At least one item is required'
                })
            
            # Parse invoice date
            try:
                invoice_date_obj = datetime.strptime(invoice_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                })
            
            # Calculate due date (30 days from invoice date)
            from datetime import timedelta
            due_date = invoice_date_obj + timedelta(days=30)
            
            # Calculate totals
            subtotal = Decimal('0.00')
            for item in items:
                qty = Decimal(str(item['quantity']))
                price = Decimal(str(item['unit_price']))
                subtotal += qty * price
            
            tax_amount = subtotal * Decimal('0.00')  # No tax for now
            total_amount = subtotal + tax_amount
            
            # Create invoice
            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                customer_name=customer_name,
                invoice_date=invoice_date_obj,
                due_date=due_date,
                subtotal=subtotal,
                tax_amount=tax_amount,
                total_amount=total_amount,
                status='draft',
                created_by=request.user
            )
            
            # Create invoice items
            for item in items:
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=item['description'],
                    quantity=int(item['quantity']),
                    unit_price=Decimal(str(item['unit_price']))
                )
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='Create Invoice',
                module='Finance',
                description=f'Created invoice {invoice.invoice_number} for {customer_name}',
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Invoice created successfully',
                'invoice': {
                    'id': invoice.id,
                    'invoice_number': invoice.invoice_number,
                    'customer_name': invoice.customer_name,
                    'total_amount': float(invoice.total_amount),
                    'status': invoice.get_status_display()
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@role_required('finance', 'admin')
def record_payment_api(request):
    """Record payment via API"""
    if request.method == 'POST':
        try:
            # Get form data
            payment_type = request.POST.get('payment_type')
            payment_method = request.POST.get('payment_method')
            amount = request.POST.get('amount')
            payment_date = request.POST.get('payment_date')
            recipient = request.POST.get('recipient')
            reference_number = request.POST.get('reference_number', '')
            notes = request.POST.get('notes', '')
            
            # Validate required fields
            if not all([payment_type, payment_method, amount, payment_date, recipient]):
                return JsonResponse({
                    'success': False,
                    'message': 'All required fields must be filled'
                })
            
            # Parse payment date
            try:
                payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                })
            
            # Generate payment number
            last_payment = Payment.objects.all().order_by('-id').first()
            if last_payment and last_payment.payment_number:
                last_num = int(last_payment.payment_number.split('-')[1])
                new_num = last_num + 1
            else:
                new_num = 1
            payment_number = f"PAY-{new_num:05d}"
            
            # Create payment
            payment = Payment.objects.create(
                payment_number=payment_number,
                payment_type=payment_type,
                payment_method=payment_method,
                amount=Decimal(amount),
                payment_date=payment_date_obj,
                recipient=recipient,
                reference_number=reference_number,
                notes=notes,
                processed_by=request.user
            )
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='Record Payment',
                module='Finance',
                description=f'Recorded {payment.get_payment_type_display()} payment of PKR {amount} to {recipient}',
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Payment recorded successfully',
                'payment': {
                    'id': payment.id,
                    'payment_number': payment.payment_number,
                    'payment_type': payment.get_payment_type_display(),
                    'amount': float(payment.amount),
                    'recipient': payment.recipient
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@role_required('finance', 'admin')
def generate_report_api(request):
    """Generate financial report via API"""
    if request.method == 'POST':
        try:
            report_type = request.POST.get('report_type')
            
            if not report_type:
                return JsonResponse({
                    'success': False,
                    'message': 'Report type is required'
                })
            
            # Get current month and year for report period
            today = date.today()
            report_period = f"{today.strftime('%B %Y')}"
            
            # Create report record
            report = FinancialReport.objects.create(
                report_type=report_type,
                report_period=report_period,
                generated_by=request.user
            )
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='Generate Report',
                module='Finance',
                description=f'Generated {report.get_report_type_display()} for {report_period}',
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'success': True,
                'message': f'{report.get_report_type_display()} generated successfully',
                'report': {
                    'id': report.id,
                    'report_type': report.get_report_type_display(),
                    'report_period': report.report_period
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@role_required('finance', 'admin')
def financial_summary_report(request):
    """Generate comprehensive financial summary report"""
    print("=" * 50)
    print("FINANCIAL SUMMARY REPORT VIEW CALLED")
    print(f"User: {request.user}")
    print(f"User Role: {request.user.role if request.user.is_authenticated else 'Not authenticated'}")
    print("=" * 50)
    
    from django.db.models import Sum, Count
    from datetime import date, timedelta
    from hr.models import Payroll
    from suppliers.models import PurchaseOrder
    
    today = date.today()
    current_month = today.month
    current_year = today.year
    
    # Calculate revenue
    total_revenue = Invoice.objects.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or 0
    monthly_revenue = Invoice.objects.filter(
        status='paid',
        invoice_date__month=current_month,
        invoice_date__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Calculate expenses from multiple sources
    # 1. Regular expenses
    expense_total = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    monthly_expense_total = Expense.objects.filter(
        expense_date__month=current_month,
        expense_date__year=current_year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # 2. Payroll expenses
    payroll_total = Payroll.objects.aggregate(total=Sum('net_pay'))['total'] or 0
    monthly_payroll_total = Payroll.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(total=Sum('net_pay'))['total'] or 0
    
    # 3. Purchase orders
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(total=Sum('total_amount'))['total'] or 0
    monthly_po_total = PurchaseOrder.objects.filter(
        status='received',
        order_date__month=current_month,
        order_date__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Total expenses
    total_expenses = expense_total + payroll_total + po_total
    monthly_expenses = monthly_expense_total + monthly_payroll_total + monthly_po_total
    
    # Calculate payments
    total_payments = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    monthly_payments = Payment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Payment breakdown by type
    payment_breakdown = Payment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).values('payment_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Expense breakdown by category - combine all sources
    expense_breakdown = []
    
    # Add regular expenses by category
    regular_expenses = Expense.objects.values('category').annotate(total=Sum('amount'))
    for exp in regular_expenses:
        expense_breakdown.append({
            'category': exp['category'],
            'total': float(exp['total'])
        })
    
    # Add payroll as labor costs
    if payroll_total > 0:
        expense_breakdown.append({
            'category': 'Labor Costs',
            'total': float(payroll_total)
        })
    
    # Add purchase orders as raw materials
    if po_total > 0:
        expense_breakdown.append({
            'category': 'Purchase Orders',
            'total': float(po_total)
        })
    
    # Calculate percentages
    if total_expenses > 0:
        for exp in expense_breakdown:
            exp['percentage'] = (exp['total'] / float(total_expenses)) * 100
    
    # Sort by total descending
    expense_breakdown = sorted(expense_breakdown, key=lambda x: x['total'], reverse=True)
    
    # Recent expenses - combine from all sources
    recent_expenses_list = []
    
    # Add regular expenses
    for exp in Expense.objects.all().order_by('-expense_date')[:5]:
        recent_expenses_list.append({
            'description': exp.description,
            'amount': float(exp.amount),
            'expense_date': exp.expense_date,
            'category': exp.get_category_display()
        })
    
    # Add recent payroll
    for payroll in Payroll.objects.all().order_by('-payment_date')[:3]:
        recent_expenses_list.append({
            'description': f'Payroll - {payroll.employee.first_name} {payroll.employee.last_name}',
            'amount': float(payroll.net_pay),
            'expense_date': payroll.payment_date,
            'category': 'Payroll'
        })
    
    # Add recent purchase orders
    for po in PurchaseOrder.objects.filter(status='received').order_by('-order_date')[:3]:
        recent_expenses_list.append({
            'description': f'Purchase Order - {po.get_supplier_name()}',
            'amount': float(po.total_amount),
            'expense_date': po.order_date,
            'category': 'Purchase Order'
        })
    
    # Sort by date and take top 5
    recent_expenses_list = sorted(recent_expenses_list, key=lambda x: x['expense_date'], reverse=True)[:5]
    
    context = {
        'report_date': today,
        'report_period': f"{today.strftime('%B %Y')}",
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'total_expenses': total_expenses,
        'monthly_expenses': monthly_expenses,
        'total_payments': total_payments,
        'monthly_payments': monthly_payments,
        'net_profit': monthly_revenue - monthly_expenses,
        'payment_breakdown': payment_breakdown,
        'expense_breakdown': expense_breakdown,
        'recent_invoices': Invoice.objects.all().order_by('-invoice_date')[:10],
        'recent_payments': Payment.objects.all().order_by('-payment_date')[:10],
        'recent_expenses': recent_expenses_list,
    }
    
    return render(request, 'finance/financial_summary_report.html', context)


@login_required
@role_required('finance', 'admin')
def dashboard_api(request):
    """Get dashboard data via API"""
    from datetime import timedelta
    from django.utils import timezone
    
    # Get current month data
    current_month = timezone.now().month
    current_year = timezone.now().year
    last_month = current_month - 1 if current_month > 1 else 12
    last_month_year = current_year if current_month > 1 else current_year - 1
    
    # Calculate total revenue (paid invoices)
    total_revenue = Invoice.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    current_month_revenue = Invoice.objects.filter(
        status='paid',
        invoice_date__month=current_month,
        invoice_date__year=current_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    last_month_revenue = Invoice.objects.filter(
        status='paid',
        invoice_date__month=last_month,
        invoice_date__year=last_month_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate revenue change percentage
    if last_month_revenue > 0:
        revenue_change = ((current_month_revenue - last_month_revenue) / last_month_revenue) * 100
    else:
        revenue_change = 100 if current_month_revenue > 0 else 0
    
    # Calculate total expenses from multiple sources
    # 1. Regular expenses from Expense model
    expense_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    expense_current_month = Expense.objects.filter(
        expense_date__month=current_month,
        expense_date__year=current_year
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    expense_last_month = Expense.objects.filter(
        expense_date__month=last_month,
        expense_date__year=last_month_year
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # 2. Payroll expenses from HR module
    payroll_total = Payroll.objects.aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    payroll_current_month = Payroll.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year
    ).aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    payroll_last_month = Payroll.objects.filter(
        payment_date__month=last_month,
        payment_date__year=last_month_year
    ).aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    
    # 3. Purchase Order expenses from Suppliers module
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    po_current_month = PurchaseOrder.objects.filter(
        status='received',
        order_date__month=current_month,
        order_date__year=current_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    po_last_month = PurchaseOrder.objects.filter(
        status='received',
        order_date__month=last_month,
        order_date__year=last_month_year
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Combine all expenses
    total_expenses = expense_total + payroll_total + po_total
    current_month_expenses = expense_current_month + payroll_current_month + po_current_month
    last_month_expenses = expense_last_month + payroll_last_month + po_last_month
    
    # Calculate expense change percentage
    if last_month_expenses > 0:
        expense_change = ((current_month_expenses - last_month_expenses) / last_month_expenses) * 100
    else:
        expense_change = 100 if current_month_expenses > 0 else 0
    
    # Calculate net profit
    net_profit = total_revenue - total_expenses
    current_month_profit = current_month_revenue - current_month_expenses
    last_month_profit = last_month_revenue - last_month_expenses
    
    # Calculate profit change percentage
    if last_month_profit != 0:
        profit_change = ((current_month_profit - last_month_profit) / abs(last_month_profit)) * 100
    else:
        profit_change = 100 if current_month_profit > 0 else 0
    
    # Pending invoices
    pending_invoices_count = Invoice.objects.filter(status__in=['draft', 'sent']).count()
    pending_invoices_amount = Invoice.objects.filter(status__in=['draft', 'sent']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Expense breakdown by category - combine all sources
    expense_breakdown = []
    
    # Add regular expenses by category
    regular_expenses = Expense.objects.values('category').annotate(total=Sum('amount'))
    for exp in regular_expenses:
        expense_breakdown.append({
            'category': exp['category'],
            'total': float(exp['total'])
        })
    
    # Add payroll as labor costs
    if payroll_total > 0:
        expense_breakdown.append({
            'category': 'labor_costs',
            'total': float(payroll_total)
        })
    
    # Add purchase orders as raw materials
    if po_total > 0:
        expense_breakdown.append({
            'category': 'raw_materials',
            'total': float(po_total)
        })
    
    # Calculate percentages
    if total_expenses > 0:
        for exp in expense_breakdown:
            exp['percentage'] = (exp['total'] / float(total_expenses)) * 100
    
    # Sort by total descending
    expense_breakdown = sorted(expense_breakdown, key=lambda x: x['total'], reverse=True)
    
    # Recent expenses - combine from all sources
    recent_expenses_list = []
    
    # Add regular expenses
    for exp in Expense.objects.all().order_by('-expense_date')[:5]:
        recent_expenses_list.append({
            'description': exp.description,
            'amount': float(exp.amount),
            'expense_date': exp.expense_date,
            'category': exp.get_category_display()
        })
    
    # Add recent payroll
    for payroll in Payroll.objects.all().order_by('-payment_date')[:3]:
        recent_expenses_list.append({
            'description': f'Payroll - {payroll.employee.first_name} {payroll.employee.last_name}',
            'amount': float(payroll.net_pay),
            'expense_date': payroll.payment_date,
            'category': 'Payroll'
        })
    
    # Add recent purchase orders
    for po in PurchaseOrder.objects.filter(status='received').order_by('-order_date')[:3]:
        recent_expenses_list.append({
            'description': f'Purchase Order - {po.get_supplier_name()}',
            'amount': float(po.total_amount),
            'expense_date': po.order_date,
            'category': 'Purchase Order'
        })
    
    # Sort by date and take top 5
    recent_expenses_list = sorted(recent_expenses_list, key=lambda x: x['expense_date'], reverse=True)[:5]
    


@login_required
@role_required('finance', 'admin')
def update_invoice_status_api(request, invoice_id):
    """Update invoice status via API"""
    if request.method == 'POST':
        try:
            invoice = get_object_or_404(Invoice, id=invoice_id)
            new_status = request.POST.get('status')
            
            if new_status not in dict(Invoice.STATUS_CHOICES):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid status'
                })
            
            old_status = invoice.status
            invoice.status = new_status
            invoice.save()
            
            # If invoice is marked as paid, automatically create a payment record
            if new_status == 'paid' and old_status != 'paid':
                # Generate payment number
                last_payment = Payment.objects.all().order_by('-id').first()
                if last_payment and last_payment.payment_number:
                    last_num = int(last_payment.payment_number.split('-')[1])
                    new_num = last_num + 1
                else:
                    new_num = 1
                payment_number = f"PAY-{new_num:05d}"
                
                # Create payment record
                from django.utils import timezone
                Payment.objects.create(
                    payment_number=payment_number,
                    payment_type='other',  # Customer payment
                    payment_method='bank_transfer',  # Default method
                    amount=invoice.total_amount,
                    payment_date=timezone.now().date(),
                    recipient=invoice.customer_name,
                    reference_number=invoice.invoice_number,
                    notes=f'Payment received for invoice {invoice.invoice_number}',
                    processed_by=request.user
                )
                
                # Log payment creation
                AuditLog.objects.create(
                    user=request.user,
                    action='Auto-Create Payment',
                    module='Finance',
                    description=f'Auto-created payment {payment_number} for invoice {invoice.invoice_number} - {invoice.customer_name}',
                    ip_address=get_client_ip(request)
                )
            
            # Log the status change
            AuditLog.objects.create(
                user=request.user,
                action='Update Invoice Status',
                module='Finance',
                description=f'Changed invoice {invoice.invoice_number} status from {old_status} to {new_status}',
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Invoice status updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@role_required('finance', 'admin')
def generate_invoice_pdf(request, invoice_id):
    """Generate PDF for invoice"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    items = InvoiceItem.objects.filter(invoice=invoice)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Company Header
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Beamy Sports ERP", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Invoice Info
    invoice_info_style = ParagraphStyle(
        'InvoiceInfo',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=6
    )
    elements.append(Paragraph(f"<b>Invoice Number:</b> {invoice.invoice_number}", invoice_info_style))
    elements.append(Paragraph(f"<b>Customer:</b> {invoice.customer_name}", invoice_info_style))
    elements.append(Paragraph(f"<b>Date:</b> {invoice.invoice_date.strftime('%B %d, %Y')}", invoice_info_style))
    elements.append(Paragraph(f"<b>Due Date:</b> {invoice.due_date.strftime('%B %d, %Y')}", invoice_info_style))
    elements.append(Paragraph(f"<b>Status:</b> {invoice.get_status_display()}", invoice_info_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Items Table
    table_data = [['Description', 'Quantity', 'Unit Price', 'Total']]
    for item in items:
        table_data.append([
            item.description,
            str(item.quantity),
            f'PKR {item.unit_price:,.2f}',
            f'PKR {item.total:,.2f}'
        ])
    
    # Add totals
    table_data.append(['', '', 'Subtotal:', f'PKR {invoice.subtotal:,.2f}'])
    table_data.append(['', '', 'Tax:', f'PKR {invoice.tax_amount:,.2f}'])
    table_data.append(['', '', 'Total:', f'PKR {invoice.total_amount:,.2f}'])
    
    table = Table(table_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -4), colors.beige),
        ('GRID', (0, 0), (-1, -4), 1, colors.black),
        ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (2, -3), (-1, -1), 'RIGHT'),
        ('LINEABOVE', (2, -3), (-1, -3), 2, colors.black),
        ('LINEABOVE', (2, -1), (-1, -1), 2, colors.black),
    ]))
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return PDF
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
    
    return response


@login_required
@role_required('finance', 'admin')
def generate_financial_report_pdf(request, report_type):
    """Generate financial reports (Income Statement, Balance Sheet, Cash Flow)"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from django.utils import timezone
    from hr.models import Payroll
    from suppliers.models import PurchaseOrder
    
    # Calculate financial data
    total_revenue = Invoice.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    expense_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    payroll_total = Payroll.objects.aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_expenses = expense_total + payroll_total + po_total
    net_profit = total_revenue - total_expenses
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    if report_type == 'income_statement':
        elements.append(Paragraph("Income Statement", title_style))
        elements.append(Paragraph(f"Beamy Sports ERP - As of {timezone.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Income Statement Table
        data = [
            ['Description', 'Amount (PKR)'],
            ['Revenue', ''],
            ['Total Revenue from Invoices', f'{total_revenue:,.2f}'],
            ['', ''],
            ['Expenses', ''],
            ['Regular Expenses', f'{expense_total:,.2f}'],
            ['Payroll Expenses', f'{payroll_total:,.2f}'],
            ['Purchase Orders', f'{po_total:,.2f}'],
            ['Total Expenses', f'{total_expenses:,.2f}'],
            ['', ''],
            ['Net Profit', f'{net_profit:,.2f}'],
        ]
        
    elif report_type == 'balance_sheet':
        elements.append(Paragraph("Balance Sheet", title_style))
        elements.append(Paragraph(f"Beamy Sports ERP - As of {timezone.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Balance Sheet Table
        data = [
            ['Description', 'Amount (PKR)'],
            ['Assets', ''],
            ['Cash and Equivalents', f'{net_profit:,.2f}'],
            ['Accounts Receivable', f'{Invoice.objects.filter(status__in=["draft", "sent"]).aggregate(Sum("total_amount"))["total_amount__sum"] or 0:,.2f}'],
            ['Total Assets', f'{net_profit + (Invoice.objects.filter(status__in=["draft", "sent"]).aggregate(Sum("total_amount"))["total_amount__sum"] or 0):,.2f}'],
            ['', ''],
            ['Liabilities', ''],
            ['Accounts Payable', f'{total_expenses:,.2f}'],
            ['Total Liabilities', f'{total_expenses:,.2f}'],
            ['', ''],
            ['Equity', ''],
            ['Retained Earnings', f'{net_profit:,.2f}'],
        ]
        
    else:  # cash_flow
        elements.append(Paragraph("Cash Flow Statement", title_style))
        elements.append(Paragraph(f"Beamy Sports ERP - As of {timezone.now().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Cash Flow Table
        data = [
            ['Description', 'Amount (PKR)'],
            ['Operating Activities', ''],
            ['Cash from Customers', f'{total_revenue:,.2f}'],
            ['Cash paid for Expenses', f'-{expense_total:,.2f}'],
            ['Cash paid for Payroll', f'-{payroll_total:,.2f}'],
            ['Cash paid for Purchases', f'-{po_total:,.2f}'],
            ['Net Cash from Operations', f'{net_profit:,.2f}'],
            ['', ''],
            ['Investing Activities', '0.00'],
            ['Financing Activities', '0.00'],
            ['', ''],
            ['Net Change in Cash', f'{net_profit:,.2f}'],
        ]
    
    table = Table(data, colWidths=[4*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return PDF
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response


@login_required
@role_required('finance', 'admin')
def export_financial_summary_pdf(request):
    """Export comprehensive financial summary to PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    from django.utils import timezone
    from hr.models import Payroll
    from suppliers.models import PurchaseOrder
    
    # Calculate financial data
    total_revenue = Invoice.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    expense_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    payroll_total = Payroll.objects.aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_expenses = expense_total + payroll_total + po_total
    net_profit = total_revenue - total_expenses
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Comprehensive Financial Summary", title_style))
    elements.append(Paragraph(f"Beamy Sports ERP - Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary Section
    summary_data = [
        ['Financial Overview', 'Amount (PKR)'],
        ['Total Revenue', f'{total_revenue:,.2f}'],
        ['Total Expenses', f'{total_expenses:,.2f}'],
        ['Net Profit/Loss', f'{net_profit:,.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey if net_profit >= 0 else colors.HexColor('#fee2e2')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Invoices Section
    elements.append(Paragraph("Recent Invoices", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    invoice_data = [['Invoice #', 'Customer', 'Date', 'Amount', 'Status']]
    for invoice in Invoice.objects.all().order_by('-invoice_date')[:10]:
        invoice_data.append([
            invoice.invoice_number,
            invoice.customer_name[:30],
            invoice.invoice_date.strftime('%Y-%m-%d'),
            f'PKR {invoice.total_amount:,.2f}',
            invoice.get_status_display()
        ])
    
    invoice_table = Table(invoice_data, colWidths=[1.5*inch, 2.5*inch, 1.2*inch, 1.5*inch, 1.2*inch])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
    ]))
    elements.append(invoice_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Expenses Breakdown
    elements.append(Paragraph("Expense Breakdown", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    expense_data = [['Category', 'Amount (PKR)', 'Percentage']]
    expense_data.append(['Regular Expenses', f'{expense_total:,.2f}', f'{(expense_total/total_expenses*100) if total_expenses > 0 else 0:.1f}%'])
    expense_data.append(['Payroll', f'{payroll_total:,.2f}', f'{(payroll_total/total_expenses*100) if total_expenses > 0 else 0:.1f}%'])
    expense_data.append(['Purchase Orders', f'{po_total:,.2f}', f'{(po_total/total_expenses*100) if total_expenses > 0 else 0:.1f}%'])
    expense_data.append(['Total', f'{total_expenses:,.2f}', '100%'])
    
    expense_table = Table(expense_data, colWidths=[3*inch, 2*inch, 2*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    elements.append(expense_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Return PDF
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_summary_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@login_required
@role_required('finance', 'admin')
def export_financial_excel(request):
    """Export comprehensive financial data to Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone
    from hr.models import Payroll
    from suppliers.models import PurchaseOrder
    from io import BytesIO
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    title_font = Font(bold=True, size=14, color='1e40af')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.create_sheet('Financial Summary')
    
    # Title
    ws_summary['A1'] = 'Financial Summary Report'
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f'Generated on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}'
    
    # Calculate totals
    total_revenue = Invoice.objects.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    expense_total = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    payroll_total = Payroll.objects.aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    po_total = PurchaseOrder.objects.filter(status='received').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_expenses = expense_total + payroll_total + po_total
    net_profit = total_revenue - total_expenses
    
    # Summary data
    ws_summary['A4'] = 'Metric'
    ws_summary['B4'] = 'Amount (PKR)'
    ws_summary['A4'].font = header_font
    ws_summary['B4'].font = header_font
    ws_summary['A4'].fill = header_fill
    ws_summary['B4'].fill = header_fill
    
    summary_data = [
        ['Total Revenue', float(total_revenue)],
        ['Regular Expenses', float(expense_total)],
        ['Payroll Expenses', float(payroll_total)],
        ['Purchase Order Expenses', float(po_total)],
        ['Total Expenses', float(total_expenses)],
        ['Net Profit/Loss', float(net_profit)],
    ]
    
    row = 5
    for item in summary_data:
        ws_summary[f'A{row}'] = item[0]
        ws_summary[f'B{row}'] = item[1]
        ws_summary[f'B{row}'].number_format = '#,##0.00'
        if 'Net Profit' in item[0]:
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'].font = Font(bold=True)
        row += 1
    
    # Auto-adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # ===== INVOICES SHEET =====
    ws_invoices = wb.create_sheet('Invoices')
    
    # Headers
    invoice_headers = ['Invoice #', 'Customer', 'Date', 'Due Date', 'Amount', 'Status']
    for col, header in enumerate(invoice_headers, 1):
        cell = ws_invoices.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    invoices = Invoice.objects.all().order_by('-invoice_date')
    for row, invoice in enumerate(invoices, 2):
        ws_invoices.cell(row, 1, invoice.invoice_number).border = border
        ws_invoices.cell(row, 2, invoice.customer_name).border = border
        ws_invoices.cell(row, 3, invoice.invoice_date.strftime('%Y-%m-%d')).border = border
        ws_invoices.cell(row, 4, invoice.due_date.strftime('%Y-%m-%d')).border = border
        ws_invoices.cell(row, 5, float(invoice.total_amount)).border = border
        ws_invoices.cell(row, 5).number_format = '#,##0.00'
        ws_invoices.cell(row, 6, invoice.get_status_display()).border = border
    
    # Auto-adjust columns
    for col in range(1, 7):
        ws_invoices.column_dimensions[get_column_letter(col)].width = 18
    
    # ===== EXPENSES SHEET =====
    ws_expenses = wb.create_sheet('Expenses')
    
    # Headers
    expense_headers = ['Expense #', 'Category', 'Description', 'Amount', 'Date']
    for col, header in enumerate(expense_headers, 1):
        cell = ws_expenses.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    expenses = Expense.objects.all().order_by('-expense_date')
    for row, expense in enumerate(expenses, 2):
        ws_expenses.cell(row, 1, expense.expense_number).border = border
        ws_expenses.cell(row, 2, expense.get_category_display()).border = border
        ws_expenses.cell(row, 3, expense.description).border = border
        ws_expenses.cell(row, 4, float(expense.amount)).border = border
        ws_expenses.cell(row, 4).number_format = '#,##0.00'
        ws_expenses.cell(row, 5, expense.expense_date.strftime('%Y-%m-%d')).border = border
    
    # Auto-adjust columns
    for col in range(1, 6):
        ws_expenses.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== PAYROLL SHEET =====
    ws_payroll = wb.create_sheet('Payroll')
    
    # Headers
    payroll_headers = ['Employee', 'Pay Period', 'Gross Pay', 'Deductions', 'Net Pay', 'Payment Date']
    for col, header in enumerate(payroll_headers, 1):
        cell = ws_payroll.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data
    payrolls = Payroll.objects.all().order_by('-payment_date')
    for row, payroll in enumerate(payrolls, 2):
        ws_payroll.cell(row, 1, f'{payroll.employee.first_name} {payroll.employee.last_name}').border = border
        ws_payroll.cell(row, 2, payroll.pay_period).border = border
        ws_payroll.cell(row, 3, float(payroll.gross_pay)).border = border
        ws_payroll.cell(row, 3).number_format = '#,##0.00'
        total_deductions = float(payroll.federal_tax + payroll.state_tax + payroll.social_security + payroll.medicare + payroll.other_deductions)
        ws_payroll.cell(row, 4, total_deductions).border = border
        ws_payroll.cell(row, 4).number_format = '#,##0.00'
        ws_payroll.cell(row, 5, float(payroll.net_pay)).border = border
        ws_payroll.cell(row, 5).number_format = '#,##0.00'
        ws_payroll.cell(row, 6, payroll.payment_date.strftime('%Y-%m-%d')).border = border
    
    # Auto-adjust columns
    for col in range(1, 7):
        ws_payroll.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return Excel file
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="financial_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
def get_customers_api(request):
    """Get list of unique customers from all orders"""
    try:
        from inventory.models import CustomerOrder
        from production.models import ProductionOrder
        
        # Get unique customer names from different sources
        customers_set = set()
        
        # From customer orders
        customer_orders = CustomerOrder.objects.values_list('customer_name', flat=True).distinct()
        customers_set.update(customer_orders)
        
        # From production orders
        production_orders = ProductionOrder.objects.values_list('customer_name', flat=True).distinct()
        customers_set.update(production_orders)
        
        # From existing invoices
        invoices = Invoice.objects.values_list('customer_name', flat=True).distinct()
        customers_set.update(invoices)
        
        # Remove empty strings
        customers_set.discard('')
        customers_set.discard(None)
        
        # Convert to sorted list
        customers_list = sorted(list(customers_set))
        
        return JsonResponse({
            'success': True,
            'customers': customers_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
